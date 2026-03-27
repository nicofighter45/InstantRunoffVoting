import pandas as pd
from openpyxl import load_workbook

from votant import Votant


excel_file = "test_results.xlsx"

def get_datas(*excel_file: str) -> tuple[list[str], list[Votant], int]:
    total_choices = []
    total_votants = []
    total_abstention = 0
    for file in excel_file:
        print(f"Loading data from {file}...")
        choices, votants, abstention = extract_data_from_excel(file)
        if total_choices != [] and choices != total_choices:
            raise Exception(f"Error: Choices in {file} do not match previous files. File probably doesn't have the correct formating. Expected choices: {total_choices}, but got {choices}.")
        total_choices = choices
        total_votants.extend(votants)
        total_abstention += abstention
        print(f"Data loading from {file} complete. Number of votants: {len(votants)}, Abstention: {abstention}.\n")
    print(f"Data loading for all file complete. Number of votants: {len(total_votants)}, Total abstention: {total_abstention}.")
    print("All votes have been extracted. Starting the counting process...\n\n")
    return total_choices, total_votants, total_abstention

def extract_data_from_excel(excel_file: str) -> tuple[list[str], list[Votant], int]:
    # Load the workbook
    wb = load_workbook(filename=excel_file, read_only=True, data_only=False)

    # Get the first sheet
    sheet = wb.active

    # Read all rows as strings
    data = []
    try:
        for row in sheet.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else '' for cell in row])
    except ValueError as e:
        pass

    # Convert to DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])

    # Replace 'NaN' with pd.NA
    df = df.replace('NaN', pd.NA).replace('N/A', pd.NA).replace('NULL', pd.NA)

    # remove first 3 columns (timestamp, email, name)
    df = df.iloc[:, 3:]

    # Get column titles as a string
    titles = df.columns.tolist()
    # Extract choices from titles (remove the first, and split by " - " to get the choice name)
    choices = [title.split(" - ")[1] for title in titles[1:]]

    votants = []
    abstention = 0

    for index, row in df.iterrows():
        vote = row.tolist()
        if vote[0] == "Oui":
            print(f"Extracting vote {index + 1}: {vote}")
            votants.append(Votant(vote[1:]))
        elif vote[0] == "Non":
            print(f"Extracting abstention {index + 1}: {vote}")
            abstention += 1

    return choices, votants, abstention
