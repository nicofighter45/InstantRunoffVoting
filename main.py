import pandas as pd
from openpyxl import load_workbook
from votant import Votant


excel_file = "test_results.xlsx"

# Load the workbook
wb = load_workbook(filename=excel_file, read_only=True, data_only=False)

# Get the first sheet
sheet = wb.active

# Read all rows as strings
data = []
try:
    for row in sheet.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else '' for cell in row])
except ValueError as e:
    pass

# Convert to DataFrame
df = pd.DataFrame(data[1:], columns=data[0])

# Replace 'NaN' with pd.NA
df = df.replace('NaN', pd.NA).replace('N/A', pd.NA).replace('NULL', pd.NA)

# remove first 3 columns (timestamp, email, name) and last column (empty)
df = df.iloc[:, 3:-1]

# Get column titles as a string
titles = df.columns.tolist()
# Extract choices from titles (remove the first, and split by " - " to get the choice name)
choices = [title.split(" - ")[1] for title in titles[1:]]

votants = []
abstention = 0

for index, row in df.iterrows():
    vote = row.tolist()
    print(f"Exctracting vote {index + 1}: {vote}")
    if vote[0] == "Non":
        votants.append(Votant(vote[1:]))
    elif vote[0] == "Non":
        abstention += 1


iter = 1
while True:
    votes = [0] * len(choices)
    for votant in votants:
        votes[votant.get_vote()] += 1

    max, max_index = 0, 0
    min, min_index = len(votants), 0
    for i in range(len(votes)):
        if votes[i] > max:
            max = votes[i]
            max_index = i
        if votes[i] < min:
            min = votes[i]
            min_index = i

    if max > len(votants) / 2:
        print(f"Le gagnant est {choices[max_index]} avec {max} votes sur {len(votants)} votants.")
        break
    else:
        iter += 1
        print(f"Pas de gagnant. Passage au {iter} tour. Le choix avec le moins de votes est {choices[min_index]} avec {min} votes sur {len(votants)} votants.")
        for votant in votants:
            votant.remove_choice(min_index)
            choices.pop(min_index)
